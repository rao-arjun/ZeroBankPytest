from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
import openpyxl
import logging

logging.basicConfig(filename="test.log", format='%(asctime)s: %(levelname)s: %(message)s', datefmt= '%m/%d/%Y %I: %M : %S %p', level=logging.DEBUG)

driver = webdriver.Chrome()
driver.maximize_window()
driver.get("http://opensource-demo.orangehrmlive.com")

logging.info("Starting webdriver")

print(driver.get_cookies())
driver.find_element_by_xpath("//*[@id='txtUsername']").send_keys("Admin")
driver.find_element_by_xpath("//*[@id='txtPassword']").send_keys("admin123")
driver.find_element_by_xpath("//*[@id='btnLogin']").click()

logging.info("Logging into the application")

wait = WebDriverWait(driver, 20)

wait.until(ec.visibility_of_element_located((By.TAG_NAME, "h1")))
print(driver.find_element_by_tag_name("h1").text)

pim = driver.find_element_by_id("menu_pim_viewPimModule")
employeelist = driver.find_element_by_id("menu_pim_viewEmployeeList")

actions = ActionChains(driver)
actions.move_to_element(pim).move_to_element(employeelist).click().perform()
# search for the link
# search employee in table and click on the link
print(driver.find_element_by_css_selector("table[id='resultTable']").is_displayed())

logging.debug("Check if table is present")
filePath = "employee.xlsx"

workbook = openpyxl.load_workbook(filePath)
sheet = workbook.get_sheet_by_name("employeeSearch")
logging.warning("Deprecated method being used to read sheet from excel file")

sheet = workbook.active
rows = sheet.max_row
cols = sheet.max_column

print(rows)
print(cols)

for r in range(2, rows+1):
    empname = sheet.cell(r, 1).value
    print(empname)

    found = False
    wait.until(ec.visibility_of(driver.find_element_by_xpath("//*[@id='resultTable']/tbody")))

    tablerows = driver.find_element_by_xpath("//*[@id='resultTable']/tbody").find_elements_by_tag_name("tr")
    print(len(tablerows))
    for tablerow in tablerows:
        tablecols = tablerow.find_elements_by_tag_name("td")
        print(len(tablecols))
        for i in range(1, len(tablecols)-5):
            linkText = tablecols[i].find_element_by_tag_name("a")
            print(linkText.text)
            logging.info("Clicking on searched employee link")
            if linkText.text == empname:
                found = True
                linkText.click()
                wait.until(ec.visibility_of(driver.find_element_by_xpath("//*[@id='profile-pic']/h1")))
                print(driver.find_element_by_xpath("//*[@id='profile-pic']/h1").text)
                logging.debug("Verifying employee searched is " + empname)
                driver.find_element_by_id("menu_pim_viewEmployeeList").click()
                break
        if found:
            break

    if found:
        sheet.cell(r, 2).value = "Search successful"

workbook.save(filePath)
driver.quit()
logging.info("Closing driver")
